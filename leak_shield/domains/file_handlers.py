from abc import ABC, abstractmethod
import csv
import os
import logging
import mimetypes
from typing import Optional
from PyPDF2 import PdfReader, PdfReadError
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class FileHandlerException(Exception):
    """Base exception for file handler errors"""
    pass


class UnsupportedFileTypeError(FileHandlerException):
    """Raised when file type is not supported"""
    pass


class FileReadError(FileHandlerException):
    """Raised when file cannot be read"""
    pass


class BaseFileHandler(ABC):
    @abstractmethod
    def extract_text(self, file_path: str) -> str:
        """Extract text content from file."""
        pass

    def _validate_file(self, file_path: str) -> None:
        """Validate file existence and readability"""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")
        if not os.access(file_path, os.R_OK):
            raise PermissionError(f"No read permission for file: {file_path}")


class TextFileHandler(BaseFileHandler):
    def extract_text(self, file_path: str) -> str:
        try:
            self._validate_file(file_path)
            with open(file_path, 'r', encoding='utf-8') as f:
                return f.read()
        except UnicodeDecodeError as e:
            logger.error(f"Unicode decode error in file {file_path}: {str(e)}")
            raise FileReadError(f"Error reading text file: {str(e)}")
        except Exception as e:
            logger.error(
                f"Unexpected error reading text file {file_path}: {str(e)}")
            raise FileReadError(
                f"Unexpected error reading text file: {str(e)}")


class PDFFileHandler(BaseFileHandler):
    def extract_text(self, file_path: str) -> str:
        try:
            self._validate_file(file_path)
            text = []
            with open(file_path, 'rb') as f:
                pdf = PdfReader(f)
                for page in pdf.pages:
                    text.append(page.extract_text())
            return '\n'.join(text)
        except PdfReadError as e:
            logger.error(f"PDF read error in file {file_path}: {str(e)}")
            raise FileReadError(f"Error reading PDF file: {str(e)}")
        except Exception as e:
            logger.error(
                f"Unexpected error reading PDF file {file_path}: {str(e)}")
            raise FileReadError(f"Unexpected error reading PDF file: {str(e)}")


class CSVFileHandler(BaseFileHandler):
    def extract_text(self, file_path: str) -> str:
        try:
            self._validate_file(file_path)
            text = []
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                for row in reader:
                    text.append(','.join(row))
            return '\n'.join(text)
        except csv.Error as e:
            logger.error(f"CSV read error in file {file_path}: {str(e)}")
            raise FileReadError(f"Error reading CSV file: {str(e)}")
        except Exception as e:
            logger.error(
                f"Unexpected error reading CSV file {file_path}: {str(e)}")
            raise FileReadError(f"Unexpected error reading CSV file: {str(e)}")


class ExcelFileHandler(BaseFileHandler):
    def extract_text(self, file_path: str) -> str:
        try:
            self._validate_file(file_path)
            text = []
            workbook = load_workbook(filename=file_path, read_only=True)

            for sheet in workbook.sheetnames:
                worksheet = workbook[sheet]
                text.append(f"Sheet: {sheet}")

                for row in worksheet.iter_rows(values_only=True):
                    row_text = [
                        str(cell) if cell is not None else '' for cell in row]
                    text.append('\t'.join(row_text))

            workbook.close()
            return '\n'.join(text)
        except InvalidFileException as e:
            logger.error(f"Excel read error in file {file_path}: {str(e)}")
            raise FileReadError(f"Error reading Excel file: {str(e)}")
        except Exception as e:
            logger.error(
                f"Unexpected error reading Excel file {file_path}: {str(e)}")
            raise FileReadError(
                f"Unexpected error reading Excel file: {str(e)}")


class FileHandlerFactory:
    HANDLERS = {
        'text/plain': TextFileHandler(),
        'application/pdf': PDFFileHandler(),
        'text/csv': CSVFileHandler(),
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet': ExcelFileHandler(),
    }

    EXTENSION_MAP = {
        '.txt': 'text/plain',
        '.pdf': 'application/pdf',
        '.csv': 'text/csv',
        '.xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        '.xls': 'application/vnd.ms-excel',
    }

    @classmethod
    def get_handler(cls, file_path: str) -> BaseFileHandler:
        try:
            if not mimetypes.inited:
                mimetypes.init()

            file_type, _ = mimetypes.guess_type(file_path)

            if not file_type:
                ext = os.path.splitext(file_path)[1].lower()
                file_type = cls.EXTENSION_MAP.get(ext)
                if not file_type:
                    logger.warning(
                        f"Unsupported file type for {file_path}, falling back to text handler")
                    return TextFileHandler()

            if file_type == 'application/vnd.ms-excel':
                file_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

            handler = cls.HANDLERS.get(file_type)
            if not handler:
                logger.warning(
                    f"No specific handler for MIME type {file_type}, falling back to text handler")
                return TextFileHandler()

            return handler

        except Exception as e:
            logger.error(
                f"Error determining file handler for {file_path}: {str(e)}")
            return TextFileHandler()
